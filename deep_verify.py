#!/usr/bin/env python3
"""
Deep Verification: Verify all 42 official data fields across all parsed resumes.
Compares extracted JSON values against actual resume text.
Rule: null/missing = acceptable (not failure). Only MISMATCH = failure.
"""

import json, os, sys, re
from dataclasses import dataclass, field
from collections import defaultdict

sys.path.insert(0, "/home/great/groq-resume-parser")
from groq_parser import extract_text_from_file, _parse_date, _calc_months

# === PATHS ===
RESUME_DIR = "/mnt/d/DOWNLOADS/Test Resumes-20260301T010115Z-1-001/Test Resumes"
RESULTS_DIR = "/home/great/groq-resume-parser/test-results/cycle2"
OUTPUT_DIR = "/home/great/groq-resume-parser/test-results"

RESUMES = [
    "Ahmad Qasem-Resume.pdf",
    "Ajay Eedara Resume...docx",
    "Ashok Kumar.doc",
    "Dexter .pdf",
    "Donald Belvin.docx",
    "Janaki T_Sr. Java.J2EE Dev_12+ Yrs.docx",
    "Jumoke-Adekanmi-Web-Developer-2025-03-21.pdf",
    "KRUPAKAR REDDY P.docx",
    "Kiran N. Penmetcha_s Profile.pdf",
    "KrupakarReddy_SystemP.docx",
    "Lakshman Podili_Senior .NET Developer.docx",
    "Mahesh_Bolikonda (1).pdf",
    "Mutchie.docx",
    "Narra reddy.docx",
    "PRANAY REDDY_DE_Resume.pdf",
    "Resume of Connal Jackson.doc",
    "Software_Developer_Resume.docx",
    "Vamsi Krishna DADE 2.docx",
    "VeerT.Net (1).docx",
    "Venkat_Rohit_Senior .NET Full Stack Developer (1).docx",
    "ZAMEN_ALADWANI_PROJECT MANAGER_09_01_2023.pdf",
    "Zaman S.pdf",
]

# Official 42 data fields
DATA_FIELDS = [
    (1, "Personal Details", "Full Name"),
    (2, "Personal Details", "First Name"),
    (3, "Personal Details", "Middle Name"),
    (4, "Personal Details", "Last Name"),
    (5, "Personal Details", "Email ID"),
    (6, "Personal Details", "Phone Number"),
    (7, "Personal Details", "Country Code"),
    (8, "Personal Details", "Social Media Links"),
    (9, "Overall Summary", "Current Job Role"),
    (10, "Overall Summary", "Relevant Job Titles"),
    (11, "Overall Summary", "Total Experience"),
    (12, "Overall Summary", "Summary"),
    (13, "Work Experiences", "Job Title"),
    (14, "Work Experiences", "Experience In Years"),
    (15, "Work Experiences", "Summary"),
    (16, "Work Experiences", "Company Name"),
    (17, "Work Experiences", "Location"),
    (18, "Work Experiences", "Start Date"),
    (19, "Work Experiences", "End Date"),
    (20, "Skills", "Skill Name"),
    (21, "Skills", "Skill Experience"),
    (22, "Skills", "Last Used"),
    (23, "Skills", "Relevant Skills"),
    (24, "Education", "Full Education Detail"),
    (25, "Education", "Type of Education"),
    (26, "Education", "Majors / Field of Study"),
    (27, "Education", "University / School Name"),
    (28, "Education", "Location"),
    (29, "Education", "Year Passed"),
    (30, "Certifications", "Certification Name"),
    (31, "Certifications", "Issuer Name"),
    (32, "Certifications", "Issued Year"),
    (33, "Languages", "Language Name"),
    (34, "Achievements", "Achievements"),
    (35, "Projects", "Project Name"),
    (36, "Projects", "Description"),
    (37, "Projects", "Company Worked"),
    (38, "Projects", "Role in Project"),
    (39, "Projects", "Start Date"),
    (40, "Projects", "End Date"),
    (41, "Key Responsibilities", "List of Key Responsibilities"),
    (42, "Domain", "List of Domains"),
]


@dataclass
class Check:
    field_num: int
    category: str
    field_name: str
    extracted: str       # what the parser returned
    status: str          # MATCH, PARTIAL, MISMATCH, NULL, N/A
    note: str = ""
    detail_idx: int = 0  # for array fields: which entry (0 = summary/first)


# === UTILITY FUNCTIONS ===

def is_null(val):
    """Check if a value is effectively null."""
    if val is None:
        return True
    if isinstance(val, str) and val.strip().lower() in ("", "null", "none", "n/a"):
        return True
    if isinstance(val, list) and len(val) == 0:
        return True
    return False


def find_in_text(needle, haystack):
    """Smart substring search. Handles short terms with word boundaries."""
    if not needle or not haystack:
        return False
    needle_clean = needle.strip().lower()
    haystack_lower = haystack.lower()
    # For initials like "N.", "E.", "S." — strip period and do simple check
    if len(needle_clean) <= 3 and needle_clean.endswith("."):
        return needle_clean in haystack_lower or needle_clean[:-1] in haystack_lower.split()
    # Names with special chars (C#, C++, .NET, etc.) — direct substring
    if any(c in needle_clean for c in '#.+/&'):
        return needle_clean in haystack_lower
    if len(needle_clean) <= 2:
        try:
            return bool(re.search(r'\b' + re.escape(needle_clean) + r'\b', haystack_lower))
        except re.error:
            return needle_clean in haystack_lower
    return needle_clean in haystack_lower


def word_overlap(phrase, text_lower):
    """Return (found, total, ratio) of significant words from phrase found in text."""
    if not phrase:
        return (0, 0, 0.0)
    words = [w for w in re.split(r'[\s/,\-]+', phrase.lower()) if len(w) > 2]
    if not words:
        return (0, 0, 0.0)
    found = sum(1 for w in words if w in text_lower)
    return (found, len(words), found / len(words))


def verify_date_in_text(date_str, text_lower):
    """Check if a date string appears in text. Returns (status, note)."""
    if is_null(date_str):
        return ("NULL", "")
    ds = str(date_str).strip().lower()
    if ds in ("present", "current", "till date"):
        if "present" in text_lower or "current" in text_lower or "till date" in text_lower:
            return ("MATCH", "Present/Current in text")
        return ("PARTIAL", "Present assumed")
    parsed = _parse_date(str(date_str))
    if not parsed:
        return ("PARTIAL", "Could not parse date")
    year, month = parsed
    year_str = str(year)
    import calendar
    month_name = calendar.month_name[month].lower() if 1 <= month <= 12 else ""
    month_abbr = calendar.month_abbr[month].lower() if 1 <= month <= 12 else ""
    mm_slash = f"{month:02d}/{year}"
    mm_dash = f"{month:02d}-{year}"
    if year_str in text_lower and (month_name in text_lower or month_abbr in text_lower
                                    or mm_slash in text_lower or mm_dash in text_lower):
        return ("MATCH", f"Date confirmed in text")
    if year_str in text_lower:
        return ("PARTIAL", f"Year {year} found, month not confirmed")
    return ("MISMATCH", f"Date {date_str} not found in text")


def truncate(val, maxlen=80):
    """Truncate string for display."""
    s = str(val)
    return s[:maxlen] + "..." if len(s) > maxlen else s


# === FIELD EXTRACTION & VERIFICATION ===

def extract_and_verify(result, text, text_lower):
    """Extract and verify all 42 fields. Returns list of Check objects."""
    checks = []
    pd = result.get("PersonalDetails") or {}
    os_ = result.get("OverallSummary") or {}
    exps = result.get("ListOfExperiences") or []
    skills = result.get("ListOfSkills") or []
    edu = result.get("ListOfEducation") or []
    certs = result.get("Certifications") or []
    langs = result.get("Languages") or []
    achv = result.get("Achievements") or []
    projs = result.get("Projects") or []

    # --- 1. Full Name ---
    name = pd.get("FullName")
    if is_null(name):
        checks.append(Check(1, "Personal Details", "Full Name", "null", "NULL"))
    else:
        parts = str(name).lower().split()
        found = sum(1 for p in parts if p in text_lower)
        if found == len(parts):
            checks.append(Check(1, "Personal Details", "Full Name", str(name), "MATCH", "All words in text"))
        elif found > 0:
            checks.append(Check(1, "Personal Details", "Full Name", str(name), "PARTIAL", f"{found}/{len(parts)} words"))
        else:
            checks.append(Check(1, "Personal Details", "Full Name", str(name), "MISMATCH", "Not in resume text"))

    # --- 2. First Name ---
    first = pd.get("FirstName")
    if is_null(first):
        checks.append(Check(2, "Personal Details", "First Name", "null", "NULL"))
    elif find_in_text(str(first), text):
        checks.append(Check(2, "Personal Details", "First Name", str(first), "MATCH"))
    else:
        checks.append(Check(2, "Personal Details", "First Name", str(first), "MISMATCH", "Not in text"))

    # --- 3. Middle Name ---
    middle = pd.get("MiddleName")
    if is_null(middle):
        checks.append(Check(3, "Personal Details", "Middle Name", "null", "NULL"))
    elif find_in_text(str(middle), text):
        checks.append(Check(3, "Personal Details", "Middle Name", str(middle), "MATCH"))
    else:
        checks.append(Check(3, "Personal Details", "Middle Name", str(middle), "MISMATCH", "Not in text"))

    # --- 4. Last Name ---
    last = pd.get("LastName")
    if is_null(last):
        checks.append(Check(4, "Personal Details", "Last Name", "null", "NULL"))
    elif find_in_text(str(last), text):
        checks.append(Check(4, "Personal Details", "Last Name", str(last), "MATCH"))
    else:
        checks.append(Check(4, "Personal Details", "Last Name", str(last), "MISMATCH", "Not in text"))

    # --- 5. Email ID ---
    email = pd.get("EmailID")
    if is_null(email):
        checks.append(Check(5, "Personal Details", "Email ID", "null", "NULL"))
    elif str(email).lower() in text_lower:
        checks.append(Check(5, "Personal Details", "Email ID", str(email), "MATCH"))
    else:
        checks.append(Check(5, "Personal Details", "Email ID", str(email), "MISMATCH", "Not in text"))

    # --- 6. Phone Number ---
    phone = pd.get("PhoneNumber")
    if is_null(phone):
        checks.append(Check(6, "Personal Details", "Phone Number", "null", "NULL"))
    else:
        digits = re.sub(r'\D', '', str(phone))
        text_digits = re.sub(r'\D', '', text)
        if len(digits) >= 7 and digits[-7:] in text_digits:
            checks.append(Check(6, "Personal Details", "Phone Number", str(phone), "MATCH", "Digits match"))
        else:
            checks.append(Check(6, "Personal Details", "Phone Number", str(phone), "PARTIAL", "Digits not fully confirmed"))

    # --- 7. Country Code ---
    cc = pd.get("CountryCode")
    if is_null(cc):
        checks.append(Check(7, "Personal Details", "Country Code", "null", "NULL"))
    elif str(cc).startswith("+"):
        checks.append(Check(7, "Personal Details", "Country Code", str(cc), "MATCH", "Has + prefix"))
    else:
        checks.append(Check(7, "Personal Details", "Country Code", str(cc), "MISMATCH", "Missing + prefix"))

    # --- 8. Social Media Links ---
    linkedin = pd.get("LinkedIn") or ""
    github = pd.get("GitHub") or ""
    portfolio = pd.get("Portfolio") or ""
    links = [l for l in [linkedin, github, portfolio] if l and str(l).strip()]
    if not links:
        checks.append(Check(8, "Personal Details", "Social Media Links", "null", "NULL"))
    else:
        link_str = " | ".join(links)
        found_any = any(str(l).lower() in text_lower for l in links)
        if found_any:
            checks.append(Check(8, "Personal Details", "Social Media Links", truncate(link_str), "MATCH"))
        else:
            checks.append(Check(8, "Personal Details", "Social Media Links", truncate(link_str), "PARTIAL", "Links extracted but not confirmed in text"))

    # --- 9. Current Job Role ---
    role = os_.get("CurrentJobRole")
    if is_null(role):
        checks.append(Check(9, "Overall Summary", "Current Job Role", "null", "NULL"))
    else:
        f, t, r = word_overlap(str(role), text_lower)
        if r >= 0.6:
            checks.append(Check(9, "Overall Summary", "Current Job Role", str(role), "MATCH", f"{f}/{t} words"))
        else:
            checks.append(Check(9, "Overall Summary", "Current Job Role", str(role), "PARTIAL", "Not fully confirmed"))

    # --- 10. Relevant Job Titles ---
    titles = os_.get("RelevantJobTitles")
    if is_null(titles):
        checks.append(Check(10, "Overall Summary", "Relevant Job Titles", "null", "NULL"))
    elif isinstance(titles, list) and len(titles) >= 3:
        checks.append(Check(10, "Overall Summary", "Relevant Job Titles", f"{len(titles)} titles", "MATCH", f"e.g. {titles[0]}"))
    elif isinstance(titles, list):
        checks.append(Check(10, "Overall Summary", "Relevant Job Titles", f"{len(titles)} titles", "PARTIAL", "Fewer than 3"))
    else:
        checks.append(Check(10, "Overall Summary", "Relevant Job Titles", str(titles)[:60], "PARTIAL"))

    # --- 11. Total Experience ---
    total_exp = os_.get("TotalExperience")
    if is_null(total_exp):
        checks.append(Check(11, "Overall Summary", "Total Experience", "null", "NULL"))
    else:
        # Recalculate from individual experiences
        calc_months = 0
        for exp in exps:
            s = _parse_date(exp.get("StartDate"))
            e = _parse_date(exp.get("EndDate"))
            if s and e:
                calc_months += _calc_months(s, e)
        calc_years = round(calc_months / 12, 1)
        ext_match = re.search(r'[\d.]+', str(total_exp))
        if ext_match:
            ext_years = float(ext_match.group())
            diff = abs(calc_years - ext_years)
            if diff <= 1.5:
                checks.append(Check(11, "Overall Summary", "Total Experience", str(total_exp), "MATCH", f"Calc={calc_years}y"))
            elif diff <= 3.0:
                checks.append(Check(11, "Overall Summary", "Total Experience", str(total_exp), "PARTIAL", f"Calc={calc_years}y, diff={diff:.1f}"))
            else:
                checks.append(Check(11, "Overall Summary", "Total Experience", str(total_exp), "MISMATCH", f"Calc={calc_years}y, diff={diff:.1f}"))
        else:
            checks.append(Check(11, "Overall Summary", "Total Experience", str(total_exp), "PARTIAL", "Cannot parse number"))

    # --- 12. Summary ---
    summary = os_.get("Summary")
    if is_null(summary):
        checks.append(Check(12, "Overall Summary", "Summary", "null", "NULL"))
    else:
        wc = len(str(summary).split())
        if wc >= 5:
            checks.append(Check(12, "Overall Summary", "Summary", truncate(str(summary), 60), "MATCH", f"{wc} words"))
        else:
            checks.append(Check(12, "Overall Summary", "Summary", truncate(str(summary), 60), "PARTIAL", "Very short"))

    # --- 13-19. Work Experiences (per entry, first/most recent shown in matrix) ---
    if not exps:
        for fnum, fname in [(13, "Job Title"), (14, "Experience In Years"), (15, "Summary"),
                            (16, "Company Name"), (17, "Location"), (18, "Start Date"), (19, "End Date")]:
            checks.append(Check(fnum, "Work Experiences", fname, "null", "NULL"))
    else:
        for idx, exp in enumerate(exps):
            # 13. Job Title
            jt = exp.get("JobTitle")
            if is_null(jt):
                checks.append(Check(13, "Work Experiences", "Job Title", "null", "NULL", detail_idx=idx))
            elif find_in_text(str(jt), text):
                checks.append(Check(13, "Work Experiences", "Job Title", str(jt), "MATCH", detail_idx=idx))
            else:
                f, t, r = word_overlap(str(jt), text_lower)
                if r >= 0.6:
                    checks.append(Check(13, "Work Experiences", "Job Title", str(jt), "MATCH", f"{f}/{t} words", idx))
                else:
                    checks.append(Check(13, "Work Experiences", "Job Title", str(jt), "MISMATCH", "Not in text", idx))

            # 14. Experience In Years
            ey = exp.get("ExperienceInYears")
            if is_null(ey):
                checks.append(Check(14, "Work Experiences", "Experience In Years", "null", "NULL", detail_idx=idx))
            else:
                sd_p = _parse_date(exp.get("StartDate"))
                ed_p = _parse_date(exp.get("EndDate"))
                if sd_p and ed_p:
                    cm = _calc_months(sd_p, ed_p)
                    cy = round(cm / 12, 1)
                    try:
                        ext = float(str(ey).replace("years","").strip())
                        diff = abs(cy - ext)
                        if diff <= 0.3:
                            checks.append(Check(14, "Work Experiences", "Experience In Years", str(ey), "MATCH", f"Calc={cy}", idx))
                        elif diff <= 1.0:
                            checks.append(Check(14, "Work Experiences", "Experience In Years", str(ey), "PARTIAL", f"Calc={cy}, diff={diff:.1f}", idx))
                        else:
                            checks.append(Check(14, "Work Experiences", "Experience In Years", str(ey), "MISMATCH", f"Calc={cy}, diff={diff:.1f}", idx))
                    except ValueError:
                        checks.append(Check(14, "Work Experiences", "Experience In Years", str(ey), "PARTIAL", "Not numeric", idx))
                else:
                    checks.append(Check(14, "Work Experiences", "Experience In Years", str(ey), "PARTIAL", "Dates not parseable", idx))

            # 15. Summary (KeyResponsibilities)
            resps = exp.get("KeyResponsibilities") or []
            exp_summary = exp.get("Summary")
            if resps and isinstance(resps, list) and len(resps) > 0:
                found_count = 0
                total_checked = min(len(resps), 10)
                for resp in resps[:10]:
                    if not isinstance(resp, str) or not resp.strip():
                        continue
                    resp_words = [w for w in resp.lower().split() if len(w) > 3][:5]
                    if resp_words and sum(1 for w in resp_words if w in text_lower) / len(resp_words) >= 0.6:
                        found_count += 1
                pct = round(found_count / total_checked * 100) if total_checked else 0
                if pct >= 60:
                    checks.append(Check(15, "Work Experiences", "Summary", f"{len(resps)} bullets ({pct}% confirmed)", "MATCH", "", idx))
                elif pct >= 30:
                    checks.append(Check(15, "Work Experiences", "Summary", f"{len(resps)} bullets ({pct}% confirmed)", "PARTIAL", "", idx))
                else:
                    checks.append(Check(15, "Work Experiences", "Summary", f"{len(resps)} bullets ({pct}% confirmed)", "MISMATCH", "Low text match", idx))
            elif not is_null(exp_summary):
                checks.append(Check(15, "Work Experiences", "Summary", truncate(str(exp_summary), 60), "PARTIAL", "Has summary text", idx))
            else:
                checks.append(Check(15, "Work Experiences", "Summary", "null", "NULL", detail_idx=idx))

            # 16. Company Name
            co = exp.get("CompanyName")
            if is_null(co):
                checks.append(Check(16, "Work Experiences", "Company Name", "null", "NULL", detail_idx=idx))
            elif find_in_text(str(co), text):
                checks.append(Check(16, "Work Experiences", "Company Name", str(co), "MATCH", detail_idx=idx))
            else:
                f, t, r = word_overlap(str(co), text_lower)
                if r >= 0.5:
                    checks.append(Check(16, "Work Experiences", "Company Name", str(co), "MATCH", f"Partial: {f}/{t} words", idx))
                else:
                    checks.append(Check(16, "Work Experiences", "Company Name", str(co), "MISMATCH", "Not in text", idx))

            # 17. Location
            loc = exp.get("Location")
            if is_null(loc):
                checks.append(Check(17, "Work Experiences", "Location", "null", "NULL", detail_idx=idx))
            else:
                industry_tags = {"banking", "healthcare", "entertainment", "wireless", "insurance", "telecom"}
                if any(t in str(loc).lower() for t in industry_tags):
                    checks.append(Check(17, "Work Experiences", "Location", str(loc), "MISMATCH", "Industry tag, not location", idx))
                elif find_in_text(str(loc), text):
                    checks.append(Check(17, "Work Experiences", "Location", str(loc), "MATCH", detail_idx=idx))
                else:
                    f, t, r = word_overlap(str(loc), text_lower)
                    if r >= 0.5:
                        checks.append(Check(17, "Work Experiences", "Location", str(loc), "MATCH", f"{f}/{t} words", idx))
                    else:
                        checks.append(Check(17, "Work Experiences", "Location", str(loc), "PARTIAL", "Not confirmed", idx))

            # 18. Start Date
            sd = exp.get("StartDate")
            sd_st, sd_note = verify_date_in_text(sd, text_lower)
            checks.append(Check(18, "Work Experiences", "Start Date",
                               str(sd) if not is_null(sd) else "null", sd_st, sd_note, idx))

            # 19. End Date
            ed = exp.get("EndDate")
            ed_st, ed_note = verify_date_in_text(ed, text_lower)
            checks.append(Check(19, "Work Experiences", "End Date",
                               str(ed) if not is_null(ed) else "null", ed_st, ed_note, idx))

    # --- 20-23. Skills ---
    if not skills:
        for fnum, fname in [(20, "Skill Name"), (21, "Skill Experience"),
                            (22, "Last Used"), (23, "Relevant Skills")]:
            checks.append(Check(fnum, "Skills", fname, "null", "NULL"))
    else:
        for idx, skill in enumerate(skills):
            if not isinstance(skill, dict):
                continue
            sname = skill.get("SkillName", "")

            # 20. Skill Name
            if is_null(sname):
                checks.append(Check(20, "Skills", "Skill Name", "null", "NULL", detail_idx=idx))
            elif find_in_text(str(sname), text):
                checks.append(Check(20, "Skills", "Skill Name", str(sname), "MATCH", detail_idx=idx))
            else:
                checks.append(Check(20, "Skills", "Skill Name", str(sname), "MISMATCH", "Not in resume", idx))

            # 21. Skill Experience
            sem = skill.get("SkillExperienceInMonths")
            if is_null(sem):
                checks.append(Check(21, "Skills", "Skill Experience", "null", "NULL", detail_idx=idx))
            else:
                checks.append(Check(21, "Skills", "Skill Experience", f"{sem} months", "MATCH", "Has value", idx))

            # 22. Last Used
            lu = skill.get("LastUsed")
            if is_null(lu):
                checks.append(Check(22, "Skills", "Last Used", "null", "NULL", detail_idx=idx))
            else:
                checks.append(Check(22, "Skills", "Last Used", str(lu), "MATCH", "Has value", idx))

            # 23. Relevant Skills
            rs = skill.get("RelevantSkills")
            if is_null(rs):
                checks.append(Check(23, "Skills", "Relevant Skills", "null", "NULL", detail_idx=idx))
            elif isinstance(rs, list) and len(rs) > 0:
                checks.append(Check(23, "Skills", "Relevant Skills", f"{len(rs)} items", "MATCH", detail_idx=idx))
            else:
                checks.append(Check(23, "Skills", "Relevant Skills", str(rs)[:40], "PARTIAL", detail_idx=idx))

    # --- 24-29. Education ---
    if not edu:
        for fnum, fname in [(24, "Full Education Detail"), (25, "Type of Education"),
                            (26, "Majors / Field of Study"), (27, "University / School Name"),
                            (28, "Location"), (29, "Year Passed")]:
            checks.append(Check(fnum, "Education", fname, "null", "NULL"))
    else:
        for idx, ed_entry in enumerate(edu):
            if not isinstance(ed_entry, dict):
                continue

            # 24. Full Education Detail
            degree = ed_entry.get("Degree") or ""
            inst = ed_entry.get("Institution") or ""
            full_edu = f"{degree}, {inst}".strip(", ")
            if is_null(full_edu):
                checks.append(Check(24, "Education", "Full Education Detail", "null", "NULL", detail_idx=idx))
            else:
                f, t, r = word_overlap(full_edu, text_lower)
                if r >= 0.5:
                    checks.append(Check(24, "Education", "Full Education Detail", truncate(full_edu), "MATCH", f"{f}/{t} words", idx))
                else:
                    checks.append(Check(24, "Education", "Full Education Detail", truncate(full_edu), "PARTIAL", f"{f}/{t} words", idx))

            # 25. Type of Education
            toe = ed_entry.get("TypeOfEducation")
            if is_null(toe):
                checks.append(Check(25, "Education", "Type of Education", "null", "NULL", detail_idx=idx))
            else:
                valid = {"full-time", "part-time", "online", "distance"}
                if str(toe).lower() in valid:
                    checks.append(Check(25, "Education", "Type of Education", str(toe), "MATCH", detail_idx=idx))
                else:
                    checks.append(Check(25, "Education", "Type of Education", str(toe), "PARTIAL", "Non-standard type", idx))

            # 26. Majors / Field of Study
            fld = ed_entry.get("Field")
            if is_null(fld):
                checks.append(Check(26, "Education", "Majors / Field of Study", "null", "NULL", detail_idx=idx))
            elif find_in_text(str(fld), text):
                checks.append(Check(26, "Education", "Majors / Field of Study", str(fld), "MATCH", detail_idx=idx))
            else:
                f, t, r = word_overlap(str(fld), text_lower)
                if r >= 0.5:
                    checks.append(Check(26, "Education", "Majors / Field of Study", str(fld), "MATCH", f"{f}/{t} words", idx))
                else:
                    checks.append(Check(26, "Education", "Majors / Field of Study", str(fld), "PARTIAL", "Not confirmed", idx))

            # 27. University / School Name
            uni = ed_entry.get("Institution")
            if is_null(uni):
                checks.append(Check(27, "Education", "University / School Name", "null", "NULL", detail_idx=idx))
            elif find_in_text(str(uni), text):
                checks.append(Check(27, "Education", "University / School Name", str(uni), "MATCH", detail_idx=idx))
            else:
                f, t, r = word_overlap(str(uni), text_lower)
                if r >= 0.5:
                    checks.append(Check(27, "Education", "University / School Name", str(uni), "MATCH", f"{f}/{t} words", idx))
                else:
                    checks.append(Check(27, "Education", "University / School Name", str(uni), "PARTIAL", "Not confirmed", idx))

            # 28. Education Location
            eloc = ed_entry.get("Location")
            if is_null(eloc):
                checks.append(Check(28, "Education", "Location", "null", "NULL", detail_idx=idx))
            elif find_in_text(str(eloc), text):
                checks.append(Check(28, "Education", "Location", str(eloc), "MATCH", detail_idx=idx))
            else:
                checks.append(Check(28, "Education", "Location", str(eloc), "PARTIAL", "Not confirmed", idx))

            # 29. Year Passed
            yp = ed_entry.get("YearPassed")
            if is_null(yp):
                checks.append(Check(29, "Education", "Year Passed", "null", "NULL", detail_idx=idx))
            elif str(yp).strip() in text:
                checks.append(Check(29, "Education", "Year Passed", str(yp), "MATCH", detail_idx=idx))
            else:
                checks.append(Check(29, "Education", "Year Passed", str(yp), "PARTIAL", "Year not confirmed", idx))

    # --- 30-32. Certifications ---
    if not certs:
        for fnum, fname in [(30, "Certification Name"), (31, "Issuer Name"), (32, "Issued Year")]:
            checks.append(Check(fnum, "Certifications", fname, "null", "NULL"))
    else:
        for idx, cert in enumerate(certs):
            if not isinstance(cert, dict):
                continue
            cn = cert.get("CertificationName") or cert.get("Name")
            if is_null(cn):
                checks.append(Check(30, "Certifications", "Certification Name", "null", "NULL", detail_idx=idx))
            else:
                f, t, r = word_overlap(str(cn), text_lower)
                if r >= 0.5:
                    checks.append(Check(30, "Certifications", "Certification Name", str(cn), "MATCH", f"{f}/{t} words", idx))
                else:
                    checks.append(Check(30, "Certifications", "Certification Name", str(cn), "PARTIAL", "Not fully confirmed", idx))

            issuer = cert.get("IssuerName")
            if is_null(issuer):
                checks.append(Check(31, "Certifications", "Issuer Name", "null", "NULL", detail_idx=idx))
            elif find_in_text(str(issuer), text):
                checks.append(Check(31, "Certifications", "Issuer Name", str(issuer), "MATCH", detail_idx=idx))
            else:
                checks.append(Check(31, "Certifications", "Issuer Name", str(issuer), "PARTIAL", "May be inferred", idx))

            iy = cert.get("IssuedYear")
            if is_null(iy):
                checks.append(Check(32, "Certifications", "Issued Year", "null", "NULL", detail_idx=idx))
            elif str(iy).strip() in text:
                checks.append(Check(32, "Certifications", "Issued Year", str(iy), "MATCH", detail_idx=idx))
            else:
                checks.append(Check(32, "Certifications", "Issued Year", str(iy), "PARTIAL", "Year not confirmed", idx))

    # --- 33. Languages ---
    if not langs:
        checks.append(Check(33, "Languages", "Language Name", "null", "NULL"))
    else:
        for idx, lang in enumerate(langs):
            if isinstance(lang, dict):
                ln = lang.get("LanguageName") or lang.get("Language") or lang.get("Name") or ""
            else:
                ln = str(lang)
            if is_null(ln):
                checks.append(Check(33, "Languages", "Language Name", "null", "NULL", detail_idx=idx))
            elif find_in_text(ln, text):
                checks.append(Check(33, "Languages", "Language Name", ln, "MATCH", "In text", idx))
            else:
                checks.append(Check(33, "Languages", "Language Name", ln, "MISMATCH", "Not in text - hallucinated", idx))

    # --- 34. Achievements ---
    if not achv:
        checks.append(Check(34, "Achievements", "Achievements", "null", "NULL"))
    else:
        for idx, a in enumerate(achv):
            if isinstance(a, dict):
                desc = a.get("Description") or a.get("Achievement") or ""
            else:
                desc = str(a)
            if is_null(desc):
                checks.append(Check(34, "Achievements", "Achievements", "null", "NULL", detail_idx=idx))
            else:
                f, t, r = word_overlap(str(desc), text_lower)
                if r >= 0.5:
                    checks.append(Check(34, "Achievements", "Achievements", truncate(str(desc), 60), "MATCH", f"{f}/{t} words", idx))
                else:
                    checks.append(Check(34, "Achievements", "Achievements", truncate(str(desc), 60), "PARTIAL", "Not fully confirmed", idx))

    # --- 35-40. Projects ---
    if not projs:
        for fnum, fname in [(35, "Project Name"), (36, "Description"), (37, "Company Worked"),
                            (38, "Role in Project"), (39, "Start Date"), (40, "End Date")]:
            checks.append(Check(fnum, "Projects", fname, "null", "NULL"))
    else:
        for idx, proj in enumerate(projs):
            if not isinstance(proj, dict):
                continue

            pname = proj.get("ProjectName")
            if is_null(pname):
                checks.append(Check(35, "Projects", "Project Name", "null", "NULL", detail_idx=idx))
            else:
                f, t, r = word_overlap(str(pname), text_lower)
                if r >= 0.5:
                    checks.append(Check(35, "Projects", "Project Name", str(pname), "MATCH", f"{f}/{t} words", idx))
                else:
                    checks.append(Check(35, "Projects", "Project Name", str(pname), "PARTIAL", "Not confirmed", idx))

            pdesc = proj.get("Description")
            if is_null(pdesc):
                checks.append(Check(36, "Projects", "Description", "null", "NULL", detail_idx=idx))
            else:
                checks.append(Check(36, "Projects", "Description", truncate(str(pdesc), 50), "MATCH", "Has content", idx))

            pco = proj.get("CompanyWorked")
            if is_null(pco):
                checks.append(Check(37, "Projects", "Company Worked", "null", "NULL", detail_idx=idx))
            elif find_in_text(str(pco), text):
                checks.append(Check(37, "Projects", "Company Worked", str(pco), "MATCH", detail_idx=idx))
            else:
                checks.append(Check(37, "Projects", "Company Worked", str(pco), "PARTIAL", "Not confirmed", idx))

            prole = proj.get("RoleInProject")
            if is_null(prole):
                checks.append(Check(38, "Projects", "Role in Project", "null", "NULL", detail_idx=idx))
            else:
                checks.append(Check(38, "Projects", "Role in Project", str(prole), "MATCH", "Has value", idx))

            psd = proj.get("StartDate")
            psd_st, psd_n = verify_date_in_text(psd, text_lower)
            checks.append(Check(39, "Projects", "Start Date",
                               str(psd) if not is_null(psd) else "null", psd_st, psd_n, idx))

            ped = proj.get("EndDate")
            ped_st, ped_n = verify_date_in_text(ped, text_lower)
            checks.append(Check(40, "Projects", "End Date",
                               str(ped) if not is_null(ped) else "null", ped_st, ped_n, idx))

    # --- 41. Key Responsibilities ---
    all_resps = []
    for exp in exps:
        r = exp.get("KeyResponsibilities") or []
        if isinstance(r, list):
            all_resps.extend([x for x in r if isinstance(x, str)])
    if not all_resps:
        checks.append(Check(41, "Key Responsibilities", "List of Key Responsibilities", "null", "NULL"))
    else:
        found = 0
        total_checked = min(len(all_resps), 20)
        for resp in all_resps[:20]:
            rw = [w for w in resp.lower().split() if len(w) > 3][:5]
            if rw and sum(1 for w in rw if w in text_lower) / len(rw) >= 0.6:
                found += 1
        pct = round(found / total_checked * 100) if total_checked else 0
        if pct >= 60:
            checks.append(Check(41, "Key Responsibilities", "List of Key Responsibilities",
                               f"{len(all_resps)} total ({pct}% confirmed)", "MATCH"))
        elif pct >= 30:
            checks.append(Check(41, "Key Responsibilities", "List of Key Responsibilities",
                               f"{len(all_resps)} total ({pct}% confirmed)", "PARTIAL"))
        else:
            checks.append(Check(41, "Key Responsibilities", "List of Key Responsibilities",
                               f"{len(all_resps)} total ({pct}% confirmed)", "MISMATCH", "Low text match"))

    # --- 42. Domain ---
    domain = os_.get("Domain")
    if is_null(domain):
        checks.append(Check(42, "Domain", "List of Domains", "null", "NULL"))
    else:
        checks.append(Check(42, "Domain", "List of Domains", str(domain), "MATCH", "Has value"))

    return checks


# === EXCEL REPORT ===

def generate_excel(all_results, resume_names):
    """Generate 4-sheet Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    NULL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    NORMAL_FONT = Font(name="Calibri", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"), bottom=Side(style="thin", color="B0B0B0"),
    )

    def style_cell(ws, row, col, status=None):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if status == "MATCH":
            cell.fill = MATCH_FILL
        elif status == "PARTIAL":
            cell.fill = PARTIAL_FILL
        elif status == "MISMATCH":
            cell.fill = MISMATCH_FILL
        elif status in ("NULL", "N/A"):
            cell.fill = NULL_FILL

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def auto_width(ws, min_w=8, max_w=40):
        from openpyxl.utils import get_column_letter
        for col_cells in ws.columns:
            mx = 0
            cl = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    mx = max(mx, len(str(cell.value)))
            ws.column_dimensions[cl].width = min(max(mx + 2, min_w), max_w)

    wb = Workbook()

    # ========== SHEET 1: Data Fields Verification (matrix) ==========
    ws1 = wb.active
    ws1.title = "Data Fields Verification"

    # Headers: S.No, Category, Data Field, Resume1, Resume2, ...
    headers = ["S.No.", "Category", "Data Field"]
    for i, rname in enumerate(resume_names):
        short_name = rname.rsplit(".", 1)[0][:25]
        headers.append(f"R{i+1} - {short_name}")

    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers))

    for row_idx, (fnum, cat, fname) in enumerate(DATA_FIELDS, 2):
        ws1.cell(row=row_idx, column=1, value=fnum)
        ws1.cell(row=row_idx, column=2, value=cat)
        ws1.cell(row=row_idx, column=3, value=fname)
        for c in range(1, 4):
            style_cell(ws1, row_idx, c)

        for res_idx, (rname, checks) in enumerate(all_results):
            col = res_idx + 4
            # Find first check for this field
            field_checks = [ck for ck in checks if ck.field_num == fnum]
            if not field_checks:
                ws1.cell(row=row_idx, column=col, value="null")
                style_cell(ws1, row_idx, col, "NULL")
                continue

            # For matrix: show summary of first entry or aggregate
            ck = field_checks[0]
            # Count statuses for array fields
            if len(field_checks) > 1:
                matches = sum(1 for c in field_checks if c.status == "MATCH")
                mismatches = sum(1 for c in field_checks if c.status == "MISMATCH")
                nulls = sum(1 for c in field_checks if c.status == "NULL")
                total = len(field_checks)
                if mismatches > 0:
                    val = f"{ck.extracted} [{matches}/{total} match]"
                    status = "MISMATCH"
                elif matches > 0:
                    val = f"{ck.extracted} [{matches}/{total} match]"
                    status = "MATCH"
                elif nulls == total:
                    val = "null"
                    status = "NULL"
                else:
                    val = f"{ck.extracted} [{matches}/{total}]"
                    status = "PARTIAL"
            else:
                val = ck.extracted
                status = ck.status

            ws1.cell(row=row_idx, column=col, value=truncate(val, 40))
            style_cell(ws1, row_idx, col, status)

    auto_width(ws1)

    # ========== SHEET 2: Detailed Results ==========
    ws2 = wb.create_sheet("Detailed Results")
    det_headers = ["Resume #", "Resume Name", "S.No.", "Category", "Data Field",
                   "Entry #", "Extracted Value", "Status", "Note"]
    for c, h in enumerate(det_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(det_headers))

    drow = 2
    for res_idx, (rname, checks) in enumerate(all_results):
        for ck in checks:
            ws2.cell(row=drow, column=1, value=res_idx + 1)
            ws2.cell(row=drow, column=2, value=rname[:35])
            ws2.cell(row=drow, column=3, value=ck.field_num)
            ws2.cell(row=drow, column=4, value=ck.category)
            ws2.cell(row=drow, column=5, value=ck.field_name)
            ws2.cell(row=drow, column=6, value=ck.detail_idx if ck.detail_idx > 0 else "")
            ws2.cell(row=drow, column=7, value=truncate(ck.extracted, 80))
            ws2.cell(row=drow, column=8, value=ck.status)
            ws2.cell(row=drow, column=9, value=ck.note)
            style_cell(ws2, drow, 8, ck.status)
            for c in range(1, 10):
                if c != 8:
                    style_cell(ws2, drow, c)
            drow += 1

    auto_width(ws2)

    # ========== SHEET 3: Summary ==========
    ws3 = wb.create_sheet("Summary")
    ws3.merge_cells("A1:F1")
    ws3.cell(row=1, column=1, value="Deep Verification Summary").font = TITLE_FONT

    # Aggregate stats
    all_checks = [ck for _, checks in all_results for ck in checks]
    total = len(all_checks)
    matches = sum(1 for c in all_checks if c.status == "MATCH")
    partials = sum(1 for c in all_checks if c.status == "PARTIAL")
    mismatches = sum(1 for c in all_checks if c.status == "MISMATCH")
    nulls = sum(1 for c in all_checks if c.status == "NULL")
    na = sum(1 for c in all_checks if c.status == "N/A")

    stats = [
        ("Total Checks", total),
        ("MATCH", f"{matches} ({round(matches/total*100,1) if total else 0}%)"),
        ("PARTIAL", f"{partials} ({round(partials/total*100,1) if total else 0}%)"),
        ("MISMATCH (Failures)", f"{mismatches} ({round(mismatches/total*100,1) if total else 0}%)"),
        ("NULL (Acceptable)", f"{nulls} ({round(nulls/total*100,1) if total else 0}%)"),
        ("N/A", f"{na}"),
        ("", ""),
        ("Accuracy (MATCH only)", f"{round(matches/(total-nulls-na)*100,1) if (total-nulls-na) else 0}%"),
        ("Accuracy (MATCH+PARTIAL)", f"{round((matches+partials)/(total-nulls-na)*100,1) if (total-nulls-na) else 0}%"),
        ("Failure Rate (MISMATCH)", f"{round(mismatches/(total-nulls-na)*100,1) if (total-nulls-na) else 0}%"),
    ]

    sum_headers = ["Metric", "Value"]
    ws3.cell(row=3, column=1, value=sum_headers[0])
    ws3.cell(row=3, column=2, value=sum_headers[1])
    style_header(ws3, 3, 2)

    for i, (metric, val) in enumerate(stats, 4):
        ws3.cell(row=i, column=1, value=metric).font = NORMAL_FONT
        ws3.cell(row=i, column=2, value=val).font = NORMAL_FONT
        for c in range(1, 3):
            style_cell(ws3, i, c)

    # Per-category breakdown
    cat_row = 4 + len(stats) + 1
    ws3.cell(row=cat_row, column=1, value="Per-Category Breakdown").font = Font(bold=True, size=12)
    cat_headers = ["Category", "MATCH", "PARTIAL", "MISMATCH", "NULL", "Total"]
    for c, h in enumerate(cat_headers, 1):
        ws3.cell(row=cat_row + 1, column=c, value=h)
    style_header(ws3, cat_row + 1, len(cat_headers))

    categories = defaultdict(lambda: {"MATCH": 0, "PARTIAL": 0, "MISMATCH": 0, "NULL": 0, "N/A": 0})
    for ck in all_checks:
        categories[ck.category][ck.status] += 1

    for i, (cat, counts) in enumerate(sorted(categories.items()), cat_row + 2):
        ws3.cell(row=i, column=1, value=cat)
        ws3.cell(row=i, column=2, value=counts["MATCH"])
        ws3.cell(row=i, column=3, value=counts["PARTIAL"])
        ws3.cell(row=i, column=4, value=counts["MISMATCH"])
        ws3.cell(row=i, column=5, value=counts["NULL"])
        ws3.cell(row=i, column=6, value=sum(counts.values()))
        for c in range(1, 7):
            style_cell(ws3, i, c)

    # Per-resume breakdown
    res_row = cat_row + 2 + len(categories) + 1
    ws3.cell(row=res_row, column=1, value="Per-Resume Breakdown").font = Font(bold=True, size=12)
    res_headers = ["Resume", "MATCH", "PARTIAL", "MISMATCH", "NULL", "Total", "Accuracy %"]
    for c, h in enumerate(res_headers, 1):
        ws3.cell(row=res_row + 1, column=c, value=h)
    style_header(ws3, res_row + 1, len(res_headers))

    for i, (rname, checks) in enumerate(all_results, res_row + 2):
        m = sum(1 for c in checks if c.status == "MATCH")
        p = sum(1 for c in checks if c.status == "PARTIAL")
        mm = sum(1 for c in checks if c.status == "MISMATCH")
        n = sum(1 for c in checks if c.status == "NULL")
        t = len(checks)
        non_null = t - n
        acc = round(m / non_null * 100, 1) if non_null else 0
        ws3.cell(row=i, column=1, value=rname[:35])
        ws3.cell(row=i, column=2, value=m)
        ws3.cell(row=i, column=3, value=p)
        ws3.cell(row=i, column=4, value=mm)
        ws3.cell(row=i, column=5, value=n)
        ws3.cell(row=i, column=6, value=t)
        ws3.cell(row=i, column=7, value=f"{acc}%")
        for c in range(1, 8):
            style_cell(ws3, i, c)

    auto_width(ws3)

    # ========== SHEET 4: Mismatches Only ==========
    ws4 = wb.create_sheet("Mismatches Only")
    mm_headers = ["Resume #", "Resume Name", "Category", "Data Field", "Entry #",
                  "Extracted Value", "Note"]
    for c, h in enumerate(mm_headers, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, len(mm_headers))

    mrow = 2
    for res_idx, (rname, checks) in enumerate(all_results):
        for ck in checks:
            if ck.status == "MISMATCH":
                ws4.cell(row=mrow, column=1, value=res_idx + 1)
                ws4.cell(row=mrow, column=2, value=rname[:35])
                ws4.cell(row=mrow, column=3, value=ck.category)
                ws4.cell(row=mrow, column=4, value=ck.field_name)
                ws4.cell(row=mrow, column=5, value=ck.detail_idx if ck.detail_idx > 0 else "")
                ws4.cell(row=mrow, column=6, value=truncate(ck.extracted, 80))
                ws4.cell(row=mrow, column=7, value=ck.note)
                for c in range(1, 8):
                    style_cell(ws4, mrow, c, "MISMATCH")
                mrow += 1

    if mrow == 2:
        ws4.cell(row=2, column=1, value="No mismatches found!").font = Font(bold=True, color="006100")

    auto_width(ws4)

    # Save
    path = os.path.join(OUTPUT_DIR, "Deep_Verification_Report.xlsx")
    wb.save(path)
    return path


# === MAIN ===

def main():
    print("=" * 70)
    print("DEEP VERIFICATION — 42 Data Fields × 22 Resumes")
    print("=" * 70)

    all_results = []
    resume_names_used = []

    for i, filename in enumerate(RESUMES, 1):
        filepath = os.path.join(RESUME_DIR, filename)
        print(f"\n{'='*60}")
        print(f"[{i:02d}/{len(RESUMES)}] {filename}")
        print(f"{'='*60}")

        if not os.path.exists(filepath):
            print("  SKIP — file not found")
            continue

        # Find matching JSON result
        safe_name = f"{i:02d}-{filename.rsplit('.', 1)[0][:40]}.json".replace(" ", "_").replace("/", "_")
        json_path = os.path.join(RESULTS_DIR, safe_name)

        if not os.path.exists(json_path):
            # Try to find by prefix
            matches = [f for f in os.listdir(RESULTS_DIR) if f.startswith(f"{i:02d}-") and f.endswith(".json")]
            if matches:
                json_path = os.path.join(RESULTS_DIR, matches[0])
            else:
                print(f"  SKIP — no parsed JSON found (parse failed)")
                # Add all-NULL entry
                null_checks = [Check(fnum, cat, fname, "null", "NULL", "Parse failed")
                               for fnum, cat, fname in DATA_FIELDS]
                all_results.append((filename, null_checks))
                resume_names_used.append(filename)
                continue

        # Load parsed result
        data = json.load(open(json_path))
        result = data.get("result", data)

        if "error" in result and "PersonalDetails" not in result:
            print(f"  SKIP — parse error: {str(result.get('error',''))[:80]}")
            null_checks = [Check(fnum, cat, fname, "null", "NULL", "Parse error")
                           for fnum, cat, fname in DATA_FIELDS]
            all_results.append((filename, null_checks))
            resume_names_used.append(filename)
            continue

        # Load resume text
        resume_text = extract_text_from_file(filepath)
        text_lower = resume_text.lower()
        print(f"  Text: {len(resume_text)} chars")

        # Run verification
        checks = extract_and_verify(result, resume_text, text_lower)
        all_results.append((filename, checks))
        resume_names_used.append(filename)

        # Print summary
        m = sum(1 for c in checks if c.status == "MATCH")
        p = sum(1 for c in checks if c.status == "PARTIAL")
        mm = sum(1 for c in checks if c.status == "MISMATCH")
        n = sum(1 for c in checks if c.status == "NULL")
        total = len(checks)
        non_null = total - n
        acc = round(m / non_null * 100, 1) if non_null else 0

        print(f"  Checks: {total} | MATCH={m} PARTIAL={p} MISMATCH={mm} NULL={n}")
        print(f"  Accuracy (excl. null): {acc}%")

        # Print mismatches
        mismatches = [c for c in checks if c.status == "MISMATCH"]
        if mismatches:
            print(f"  MISMATCHES ({len(mismatches)}):")
            for c in mismatches:
                idx_str = f"[{c.detail_idx}]" if c.detail_idx > 0 else ""
                print(f"    ✗ {c.category}.{c.field_name}{idx_str} = {truncate(c.extracted, 50)} — {c.note}")

    # === FINAL SUMMARY ===
    print(f"\n\n{'='*70}")
    print("DEEP VERIFICATION — FINAL SUMMARY")
    print(f"{'='*70}")

    all_checks = [ck for _, checks in all_results for ck in checks]
    total = len(all_checks)
    matches = sum(1 for c in all_checks if c.status == "MATCH")
    partials = sum(1 for c in all_checks if c.status == "PARTIAL")
    mismatches = sum(1 for c in all_checks if c.status == "MISMATCH")
    nulls = sum(1 for c in all_checks if c.status == "NULL")
    non_null = total - nulls

    print(f"\nResumes verified: {len(all_results)}")
    print(f"Total checks:     {total}")
    print(f"  MATCH:    {matches} ({round(matches/total*100,1)}%)")
    print(f"  PARTIAL:  {partials} ({round(partials/total*100,1)}%)")
    print(f"  MISMATCH: {mismatches} ({round(mismatches/total*100,1)}%)")
    print(f"  NULL:     {nulls} ({round(nulls/total*100,1)}%)")
    print(f"\nAccuracy (excl. null):")
    print(f"  MATCH:          {round(matches/non_null*100,1) if non_null else 0}%")
    print(f"  MATCH+PARTIAL:  {round((matches+partials)/non_null*100,1) if non_null else 0}%")
    print(f"  MISMATCH rate:  {round(mismatches/non_null*100,1) if non_null else 0}%")

    # Generate Excel
    print("\nGenerating Excel report...")
    path = generate_excel(all_results, resume_names_used)
    print(f"Report saved to: {path}")


if __name__ == "__main__":
    main()
